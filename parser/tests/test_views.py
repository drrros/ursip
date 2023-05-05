from datetime import date
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from parser.models import CompanyData


class UploadFileViewTest(TestCase):
    def test_upload_file_view_uses_correct_template(self):
        response = self.client.get(reverse('index'))
        self.assertTemplateUsed(response, 'parser/index.html')

    def test_upload_file_view_uploading_file(self):
        test_file = BytesIO()
        wb = Workbook()
        ws = wb.active

        headers = ['', 'company', 'fact', '', 'forecast', '', '', '', '', '']
        for col, header in enumerate(headers):
            ws.cell(row=1, column=col + 1, value=header)

        data = [
            ['', '', 'Qliq', 'Qoil', 'Qliq', 'Qoil', '', '', '', '', '', ''],
            ['', 'data1', 'data2', 'data1', 'data2', 'data1', 'data2', 'data1', 'data2', '', '', ''],
            [1, 'company1', 10, 20, 30, 40, 12, 22, 15, 25, '', ''],
        ]

        for row, row_data in enumerate(data):
            for col, cell_data in enumerate(row_data):
                ws.cell(row=row + 2, column=col + 1, value=cell_data)

        wb.save(test_file)
        test_file.seek(0)

        test_file = SimpleUploadedFile('test_file.xlsx', test_file.getvalue(),
                                       content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response = self.client.post(reverse('index'), {'file': test_file}, follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(CompanyData.objects.filter(company='company1').exists())
        self.assertRedirects(response, reverse('results'))


class CompanyDataViewByDateTest(TestCase):

    def setUp(self):
        CompanyData.objects.create(
            id=1,
            company="company1",
            fact_qliq_data1=10,
            fact_qliq_data2=20,
            fact_qoil_data1=30,
            fact_qoil_data2=40,
            forecast_qliq_data1=10,
            forecast_qliq_data2=20,
            forecast_qoil_data1=30,
            forecast_qoil_data2=40,
            date=date(2023, 1, 1),
        )
        CompanyData.objects.create(
            id=2,
            company="company2",
            fact_qliq_data1=11,
            fact_qliq_data2=21,
            fact_qoil_data1=31,
            fact_qoil_data2=41,
            forecast_qliq_data1=10,
            forecast_qliq_data2=20,
            forecast_qoil_data1=30,
            forecast_qoil_data2=40,
            date=date(2023, 1, 1),
        )

    def test_view_uses_correct_template(self):
        response = self.client.get(reverse('totals'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'parser/results_by_date.html')

    def test_view_context_data(self):
        response = self.client.get(reverse('totals'))
        self.assertEqual(response.status_code, 200)

        totals = response.context['totals']

        self.assertEqual(len(totals), 2)
        self.assertEqual(totals[0]['Qoil_total'], 70)
        self.assertEqual(totals[0]['Qliq_total'], 30)
        self.assertEqual(totals[1]['Qoil_total'], 72)
        self.assertEqual(totals[1]['Qliq_total'], 32)

    def test_str_model(self):
        self.assertEqual(str(CompanyData.objects.get(id=1)), 'company1')